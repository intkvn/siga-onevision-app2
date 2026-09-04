import os
import tempfile
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from fastapi import APIRouter, Request, Form, Depends, UploadFile, File
from fastapi.responses import RedirectResponse, HTMLResponse, FileResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import LoteCarga, BienAlta, Pecosa
from app.auth import requiere_login
from app.services.excel_onevision import leer_reporte_qr_onevision, corregir_codigo_patrimonial
from app.services.lote_status import expedientes_de_lote, expedientes_de_lotes

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

ENCABEZADOS_BARTENDER = [
    "Codigo Patrimonial", "Codigo QR", "Ruta QR", "Bien", "Establecimiento",
    "Marca", "Modelo", "Nro Serie", "Numero pecosa",
]
ENCABEZADOS_HOJA1 = [
    "ITEM", "Codigo Patrimonial", "Codigo QR", "Bien", "Establecimiento",
    "Marca", "Modelo", "Nro Serie", "Numero pecosa",
]


@router.get("/impresion", response_class=HTMLResponse)
def formulario_impresion(request: Request, db: Session = Depends(get_db), _=Depends(requiere_login)):
    lotes_query = db.query(LoteCarga).order_by(LoteCarga.id.desc()).all()
    bienes_por_lote = defaultdict(list)
    lote_ids = [lote.id for lote in lotes_query]
    if lote_ids:
        for bien in db.query(BienAlta).filter(BienAlta.lote_id.in_(lote_ids)).all():
            bienes_por_lote[bien.lote_id].append(bien)
    expedientes_por_lote = expedientes_de_lotes(db, lotes_query)
    lotes = [
        _resumen_impresion_lote(
            db,
            lote,
            bienes=bienes_por_lote[lote.id],
            expedientes=expedientes_por_lote.get(lote.id, []),
        )
        for lote in lotes_query
    ]
    return templates.TemplateResponse("impresion.html", {"request": request, "lotes": lotes})


def _resumen_impresion_lote(
    db: Session,
    lote: LoteCarga,
    bienes: list[BienAlta] | None = None,
    expedientes: list[str] | None = None,
) -> dict:
    if bienes is None:
        bienes = db.query(BienAlta).filter(BienAlta.lote_id == lote.id).all()
    con_qr = [b for b in bienes if b.codigo_qr]

    if not bienes:
        estado = "Lote vacío (normaliza primero)"
    elif not con_qr:
        estado = "Sin reporte QR cargado"
    elif len(con_qr) < len(bienes):
        estado = f"Parcial ({len(con_qr)}/{len(bienes)})"
    else:
        estado = "Completo — listo para BarTender"

    return {
        "lote": lote,
        "estado": estado,
        "expedientes": expedientes if expedientes is not None else expedientes_de_lote(db, lote),
        "total": len(bienes),
        "con_qr": len(con_qr),
    }


@router.post("/impresion/procesar")
async def procesar_reporte_qr(
    lote_id: int = Form(...),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(requiere_login),
):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await archivo.read())
        ruta_temporal = tmp.name

    try:
        df_qr = leer_reporte_qr_onevision(ruta_temporal)
    finally:
        os.remove(ruta_temporal)

    bienes = (
        db.query(BienAlta)
        .options(joinedload(BienAlta.pecosa))
        .filter(BienAlta.lote_id == lote_id)
        .all()
    )
    bienes_por_codigo = {corregir_codigo_patrimonial(b.codigo_patrimonial): b for b in bienes}

    for _, fila in df_qr.iterrows():
        codigo = fila["codigo_patrimonial_corregido"]
        bien = bienes_por_codigo.get(codigo)
        if bien is None:
            continue  # este QR del reporte de One Visión no pertenece a este lote

        for col_qr in ["Código QR", "Codigo QR"]:
            if col_qr in fila:
                bien.codigo_qr = str(fila[col_qr])
        for col_ruta in ["Ruta QR", "URL", "Ruta"]:
            if col_ruta in fila:
                bien.ruta_qr = str(fila[col_ruta])

        if bien.pecosa:
            bien.pecosa.estado = "StickerGenerado"

    db.commit()
    return RedirectResponse(url=f"/impresion/resultado/{lote_id}", status_code=303)


@router.get("/impresion/resultado/{lote_id}", response_class=HTMLResponse)
def resultado_impresion(
    lote_id: int, request: Request, db: Session = Depends(get_db), _=Depends(requiere_login)
):
    """Muestra qué bienes del lote sí tienen su código QR (encontrados en el
    reporte de One Visión que subiste) y cuáles todavía no."""
    bienes = (
        db.query(BienAlta)
        .options(joinedload(BienAlta.pecosa))
        .filter(BienAlta.lote_id == lote_id)
        .all()
    )
    encontrados = [b for b in bienes if b.codigo_qr]
    no_encontrados = [b for b in bienes if not b.codigo_qr]
    return templates.TemplateResponse(
        "impresion_resultado.html",
        {
            "request": request, "lote_id": lote_id,
            "encontrados": encontrados, "no_encontrados": no_encontrados,
        },
    )


@router.get("/impresion/lote/{lote_id}/descargar")
def descargar_impresion(lote_id: int, db: Session = Depends(get_db), _=Depends(requiere_login)):
    bienes = (
        db.query(BienAlta)
        .options(
            joinedload(BienAlta.pecosa).joinedload(Pecosa.expediente),
            joinedload(BienAlta.centro_costo),
        )
        .filter(BienAlta.lote_id == lote_id, BienAlta.codigo_qr.isnot(None))
        .all()
    )
    expedientes = sorted({
        b.pecosa.expediente.numero for b in bienes if b.pecosa and b.pecosa.expediente
    })

    # .xlsx (no .xls) porque BarTender no acepta el formato binario antiguo.
    nombre_archivo = f"impresion_stickers_lote_{lote_id}.xlsx"
    ruta_salida = os.path.join(tempfile.gettempdir(), nombre_archivo)
    _generar_hoja_impresion(bienes, expedientes, lote_id, ruta_salida)

    return FileResponse(
        ruta_salida,
        filename=nombre_archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _generar_hoja_impresion(bienes, expedientes, lote_id, ruta_salida):
    """Genera el .xlsx con: hoja 'BarTender' (datos limpios para imprimir el
    sticker) y 'Hoja 1' (listado con encabezado de expedientes, para el
    responsable de pegar los stickers)."""
    libro = Workbook()

    hoja_bt = libro.active
    hoja_bt.title = "BarTender"
    hoja_bt.append(ENCABEZADOS_BARTENDER)
    for bien in bienes:
        hoja_bt.append(_fila_bartender(bien))

    hoja1 = libro.create_sheet("Hoja 1")
    hoja1.cell(row=1, column=1, value=f"EXPEDIENTE(S): {';'.join(expedientes)}")
    hoja1.cell(row=2, column=1, value=f"LOTE: {lote_id}")
    hoja1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ENCABEZADOS_HOJA1))
    hoja1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ENCABEZADOS_HOJA1))
    hoja1["A1"].font = Font(bold=True, size=12)
    hoja1["A2"].font = Font(bold=True, size=12)

    for col, titulo in enumerate(ENCABEZADOS_HOJA1, start=1):
        celda = hoja1.cell(row=4, column=col, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F4E78")
        celda.alignment = Alignment(horizontal="center", vertical="center")

    borde_fino = Side(style="thin", color="A6A6A6")
    borde_tabla = Border(
        left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino
    )
    for item, bien in enumerate(bienes, start=1):
        fila_idx = item + 4
        for col, valor in enumerate(_fila_hoja1(bien, item), start=1):
            celda = hoja1.cell(row=fila_idx, column=col, value=valor)
            celda.border = borde_tabla
            celda.alignment = Alignment(vertical="top")
            if col in (2, 3, 9):
                celda.number_format = "@"

    for celda in hoja1[4]:
        celda.border = borde_tabla

    anchos = [9, 22, 18, 42, 32, 18, 18, 20, 18]
    for col, ancho in enumerate(anchos, start=1):
        hoja1.column_dimensions[hoja1.cell(row=4, column=col).column_letter].width = ancho

    hoja1.auto_filter.ref = f"A4:{hoja1.cell(row=4, column=len(ENCABEZADOS_HOJA1)).column_letter}{max(4, hoja1.max_row)}"
    hoja1.freeze_panes = "A5"
    hoja1.sheet_view.showGridLines = False
    hoja1.page_setup.orientation = "landscape"
    hoja1.page_setup.fitToWidth = 1
    hoja1.page_setup.fitToHeight = 0
    hoja1.sheet_properties.pageSetUpPr.fitToPage = True
    hoja1.print_title_rows = "1:4"

    libro.save(ruta_salida)


def _fila_bartender(bien) -> list:
    return [
        bien.codigo_patrimonial,
        bien.codigo_qr or "",
        bien.ruta_qr or "",
        bien.descripcion,
        bien.centro_costo.nombre_depend if bien.centro_costo else "",
        bien.marca or "",
        bien.modelo or "",
        bien.nro_serie or "",
        bien.pecosa.numero if bien.pecosa else "",
    ]


def _fila_hoja1(bien, item: int) -> list:
    return [
        item,
        bien.codigo_patrimonial,
        bien.codigo_qr or "",
        bien.descripcion,
        bien.centro_costo.nombre_depend if bien.centro_costo else "",
        bien.marca or "",
        bien.modelo or "",
        bien.nro_serie or "",
        bien.pecosa.numero if bien.pecosa else "",
    ]
